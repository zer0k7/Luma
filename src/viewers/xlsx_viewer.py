"""Spreadsheet (.xlsx) viewer widget.

Parses OpenXML SpreadsheetML files using openpyxl and displays sheets
in a tabbed notebook containing tabular grids with headers.
"""

from pathlib import Path
from typing import Any

import gi
import openpyxl  # type: ignore

gi.require_version("Gtk", "4.0")
from gi.repository import Gtk  # type: ignore # noqa: E402

from src.strings import (  # noqa: E402
    ERROR_PARSING_FAILED,
    XLSX_EMPTY_SHEET,
    XLSX_TAB_DEFAULT_TITLE,
)
from src.viewers.base import FormatViewerError  # noqa: E402


class XlsxViewer(Gtk.Notebook):
    """Tabbed spreadsheet viewer displaying multiple worksheets in grid layouts."""

    def __init__(self, file_path: str) -> None:
        """Initialize XlsxViewer with parsed workbook sheets.

        Args:
            file_path: Absolute filesystem path to .xlsx file.

        Raises:
            FormatViewerError: If workbook parsing fails.
        """
        super().__init__()
        self.set_hexpand(True)
        self.set_vexpand(True)
        self.file_path = file_path

        self._load_workbook(file_path)

    def _load_workbook(self, file_path: str) -> None:
        """Load and render each worksheet in the workbook into a notebook tab.

        Args:
            file_path: Absolute path to the spreadsheet file.

        Raises:
            FormatViewerError: If openpyxl cannot parse the workbook.
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                tab_widget = self._create_sheet_widget(sheet)
                tab_label = Gtk.Label(
                    label=sheet_name or XLSX_TAB_DEFAULT_TITLE.format(index=idx + 1)
                )
                self.append_page(tab_widget, tab_label)
            wb.close()
        except Exception as exc:
            filename = Path(file_path).name
            error_msg = ERROR_PARSING_FAILED.format(
                path=filename,
                format_name="Excel Spreadsheet (.xlsx)",
            )
            raise FormatViewerError(error_msg, path=file_path) from exc

    def _create_sheet_widget(self, sheet: Any) -> Gtk.Widget:
        """Create a scrolled grid widget for a single worksheet.

        Args:
            sheet: Openpyxl Worksheet object.

        Returns:
            A Gtk.ScrolledWindow containing cell data.
        """
        scrolled = Gtk.ScrolledWindow()
        scrolled.set_hexpand(True)
        scrolled.set_vexpand(True)

        grid = Gtk.Grid()
        grid.set_row_spacing(4)
        grid.set_column_spacing(12)
        grid.set_margin_top(12)
        grid.set_margin_bottom(12)
        grid.set_margin_start(16)
        grid.set_margin_end(16)

        row_count = 0
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            row_count += 1
            if row_count > 500:
                break
            for col_idx, cell_value in enumerate(row):
                cell_text = str(cell_value) if cell_value is not None else ""
                lbl = Gtk.Label(label=cell_text)
                lbl.set_xalign(0.0)
                if row_idx == 0:
                    lbl.add_css_class("heading")
                grid.attach(lbl, col_idx, row_idx, 1, 1)

        if row_count == 0:
            empty_lbl = Gtk.Label(label=XLSX_EMPTY_SHEET)
            scrolled.set_child(empty_lbl)
        else:
            scrolled.set_child(grid)

        return scrolled
